# app/ui_cnc_analyzer.py
from __future__ import annotations

import hashlib
import os
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Tuple

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
except Exception:
    canvas = None
    letter = None

from .config import CNC_PROGRAMS_DIR, CNC_EXPORTS_DIR
from .db import (
    list_lines,
    list_machines,
    list_cnc_programs,
    upsert_cnc_program,
    list_cnc_program_revisions,
    latest_cnc_program_revision,
    add_cnc_program_revision,
    next_cnc_program_revision,
    get_cnc_program_revision,
    upsert_cnc_code_catalog,
    list_cnc_code_catalog,
    add_cnc_analysis_run,
    list_cnc_analysis_runs,
    add_cnc_finding,
    list_cnc_findings,
)


@dataclass
class Finding:
    severity: str
    rule_id: str
    line_numbers: List[int]
    message: str
    impact_seconds: float = 0.0


class CNCAnalyzerUI(tk.Frame):
    def __init__(self, parent, controller, show_header=False):
        super().__init__(parent, bg=controller.colors["bg"])
        self.controller = controller

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=10)

        self.tab_library = tk.Frame(nb)
        self.tab_analysis = tk.Frame(nb)
        self.tab_tool = tk.Frame(nb)
        self.tab_reports = tk.Frame(nb)
        nb.add(self.tab_library, text="Program Library")
        nb.add(self.tab_analysis, text="Analysis")
        nb.add(self.tab_tool, text="Tool Summary")
        nb.add(self.tab_reports, text="Reports")

        self._build_library()
        self._build_analysis()
        self._build_tool_summary()
        self._build_reports()

        self.current_program_id = None
        self.current_revision_id = None
        self.current_lines: List[str] = []
        self.current_findings: List[Finding] = []
        self.current_tools: Dict[str, Dict[str, float]] = {}
        self.current_efficiency = 0.0
        self.current_cycle_time = 0.0

    # ------------------- Program Library -------------------
    def _build_library(self):
        top = tk.Frame(self.tab_library)
        top.pack(fill="x")

        tk.Button(top, text="Import Program", command=self.import_program).pack(side="left")

        cols = ("program", "latest_rev", "imported_at")
        self.program_tree = ttk.Treeview(self.tab_library, columns=cols, show="headings", height=12)
        for c in cols:
            self.program_tree.heading(c, text=c.upper())
            self.program_tree.column(c, width=200)
        self.program_tree.pack(fill="x", padx=10, pady=10)
        self.program_tree.bind("<<TreeviewSelect>>", self.load_program_selection)

        viewer_frame = tk.Frame(self.tab_library)
        viewer_frame.pack(fill="both", expand=True, padx=10, pady=10)

        tk.Label(viewer_frame, text="Revision:").pack(anchor="w")
        self.revision_var = tk.StringVar(value="")
        self.revision_cb = ttk.Combobox(viewer_frame, state="readonly", textvariable=self.revision_var, width=18)
        self.revision_cb.pack(anchor="w", pady=(0, 8))
        self.revision_cb.bind("<<ComboboxSelected>>", self.load_revision_view)

        self.program_view = tk.Text(viewer_frame, height=18, wrap="none")
        self.program_view.pack(fill="both", expand=True)

        self.code_panel = tk.Listbox(viewer_frame, height=6)
        self.code_panel.pack(fill="x", pady=(8, 0))

        self.refresh_program_list()

    def refresh_program_list(self):
        for i in self.program_tree.get_children():
            self.program_tree.delete(i)
        for prog in list_cnc_programs():
            latest = latest_cnc_program_revision(prog["id"])
            self.program_tree.insert("", "end", values=(
                prog["program_name"],
                latest["revision"] if latest else "",
                latest["imported_at"] if latest else "",
            ))

    def import_program(self):
        path = filedialog.askopenfilename(
            title="Import CNC Program",
            filetypes=[("CNC Programs", "*.nc *.txt *.tap")],
        )
        if not path:
            return
        program_name = os.path.splitext(os.path.basename(path))[0]
        with open(path, "rb") as f:
            data = f.read()
        file_hash = hashlib.sha256(data).hexdigest()

        program_id = upsert_cnc_program(program_name)
        latest = latest_cnc_program_revision(program_id)
        if latest and latest.get("file_hash") == file_hash:
            messagebox.showinfo("No Change", "Program matches latest revision. No new revision created.")
            return

        revision = next_cnc_program_revision(program_id)
        dest_dir = os.path.join(CNC_PROGRAMS_DIR, program_name, f"rev{revision}")
        os.makedirs(dest_dir, exist_ok=True)
        dest_path = os.path.join(dest_dir, os.path.basename(path))
        shutil.copy2(path, dest_path)

        revision_id = add_cnc_program_revision(program_id, revision, dest_path, file_hash)
        self._update_code_catalog(revision_id, dest_path)
        self.refresh_program_list()

    def load_program_selection(self, event=None):
        sel = self.program_tree.selection()
        if not sel:
            return
        program_name = self.program_tree.item(sel[0], "values")[0]
        program = next((p for p in list_cnc_programs() if p["program_name"] == program_name), None)
        if not program:
            return
        self.current_program_id = program["id"]
        revisions = list_cnc_program_revisions(program["id"])
        self.revision_cb["values"] = [str(r["revision"]) for r in revisions]
        if revisions:
            self.revision_var.set(str(revisions[0]["revision"]))
            self._load_revision(revisions[0]["id"])

    def load_revision_view(self, event=None):
        if not self.current_program_id:
            return
        rev = self.revision_var.get()
        if not rev:
            return
        revisions = list_cnc_program_revisions(self.current_program_id)
        match = next((r for r in revisions if str(r["revision"]) == rev), None)
        if match:
            self._load_revision(match["id"])

    def _load_revision(self, revision_id: int):
        rev = get_cnc_program_revision(revision_id)
        if not rev:
            return
        self.current_revision_id = revision_id
        lines = self._read_program_lines(rev["file_path"])
        self.current_lines = lines
        self._update_viewer(lines)
        self._refresh_code_panel(revision_id)
        self._refresh_observed_codes(revision_id)
        self._refresh_analysis_history()

    def _update_viewer(self, lines: List[str]):
        self.program_view.delete("1.0", tk.END)
        for idx, line in enumerate(lines, start=1):
            self.program_view.insert(tk.END, f"{idx:>5}  {line}")

    def _read_program_lines(self, path: str) -> List[str]:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return [ln.rstrip("\n") for ln in f.readlines()]
        except Exception:
            return []

    def _update_code_catalog(self, revision_id: int, path: str):
        lines = self._read_program_lines(path)
        code_counts = {}
        sample_lines = {}
        for idx, line in enumerate(lines, start=1):
            for code in re.findall(r"\b([GM]\d+)\b", line.upper()):
                code_counts[code] = code_counts.get(code, 0) + 1
                sample_lines.setdefault(code, idx)
        for code, count in code_counts.items():
            code_type = "G" if code.startswith("G") else "M"
            upsert_cnc_code_catalog(revision_id, code, code_type, count, sample_lines.get(code, 0))

    def _refresh_code_panel(self, revision_id: int):
        self.code_panel.delete(0, "end")
        for row in list_cnc_code_catalog(revision_id):
            self.code_panel.insert(
                "end",
                f"{row['code_type']}{row['code'][1:]} | Count: {row['count']} | Sample line: {row['sample_line']}"
            )

    def _refresh_observed_codes(self, revision_id: int):
        self.observed_codes.delete(0, "end")
        codes = list_cnc_code_catalog(revision_id)
        if not codes and self.current_lines:
            rev = get_cnc_program_revision(revision_id)
            if rev:
                self._update_code_catalog(revision_id, rev["file_path"])
                codes = list_cnc_code_catalog(revision_id)
        for row in codes:
            self.observed_codes.insert(
                "end",
                f"{row['code_type']}{row['code'][1:]} | Count {row['count']}"
            )

    # ------------------- Analysis -------------------
    def _build_analysis(self):
        top = tk.Frame(self.tab_analysis)
        top.pack(fill="x", pady=10)

        tk.Label(top, text="Program:").pack(side="left")
        self.analysis_prog_var = tk.StringVar(value="")
        self.analysis_prog_cb = ttk.Combobox(top, state="readonly", textvariable=self.analysis_prog_var, width=24)
        self.analysis_prog_cb.pack(side="left", padx=6)
        self.analysis_prog_cb.bind("<<ComboboxSelected>>", self._analysis_program_changed)

        tk.Label(top, text="Revision:").pack(side="left", padx=(10, 0))
        self.analysis_rev_var = tk.StringVar(value="")
        self.analysis_rev_cb = ttk.Combobox(top, state="readonly", textvariable=self.analysis_rev_var, width=10)
        self.analysis_rev_cb.pack(side="left", padx=6)
        self.analysis_rev_cb.bind("<<ComboboxSelected>>", self._analysis_revision_changed)

        tk.Label(top, text="Machine Profile:").pack(side="left", padx=(10, 0))
        self.machine_profile_var = tk.StringVar(value="")
        profiles = [m.get("machine_number") for m in list_machines()] or list_lines()
        self.machine_profile_cb = ttk.Combobox(top, textvariable=self.machine_profile_var, values=profiles, width=18)
        self.machine_profile_cb.pack(side="left", padx=6)

        tk.Button(top, text="Run Analysis", command=self.run_analysis).pack(side="left", padx=8)

        self.analysis_summary = tk.Text(self.tab_analysis, height=8, wrap="word")
        self.analysis_summary.pack(fill="x", padx=10)

        findings_frame = tk.Frame(self.tab_analysis)
        findings_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.findings_list = tk.Listbox(findings_frame, height=10)
        self.findings_list.pack(side="left", fill="both", expand=True)
        self.findings_list.bind("<<ListboxSelect>>", self._select_finding)

        view_wrap = tk.Frame(findings_frame)
        view_wrap.pack(side="left", fill="both", expand=True, padx=8)

        self.analysis_view = tk.Text(view_wrap, height=18, wrap="none")
        self.analysis_view.pack(fill="both", expand=True)

        history_frame = tk.Frame(self.tab_analysis)
        history_frame.pack(fill="x", padx=10, pady=(0, 10))
        tk.Label(history_frame, text="Previous Runs:").pack(side="left")
        self.history_cb = ttk.Combobox(history_frame, state="readonly", width=30)
        self.history_cb.pack(side="left", padx=6)
        self.history_cb.bind("<<ComboboxSelected>>", self._load_history_run)

        self.observed_codes = tk.Listbox(self.tab_analysis, height=6)
        self.observed_codes.pack(fill="x", padx=10, pady=(0, 10))

        self._refresh_program_dropdowns()

    def _refresh_program_dropdowns(self):
        programs = list_cnc_programs()
        self.analysis_prog_cb["values"] = [p["program_name"] for p in programs]
        if programs:
            self.analysis_prog_var.set(programs[0]["program_name"])
            self._analysis_program_changed()

    def _analysis_program_changed(self, event=None):
        program_name = self.analysis_prog_var.get()
        program = next((p for p in list_cnc_programs() if p["program_name"] == program_name), None)
        if not program:
            return
        revisions = list_cnc_program_revisions(program["id"])
        self.analysis_rev_cb["values"] = [str(r["revision"]) for r in revisions]
        if revisions:
            self.analysis_rev_var.set(str(revisions[0]["revision"]))
            self._analysis_revision_changed()

    def _analysis_revision_changed(self, event=None):
        program_name = self.analysis_prog_var.get()
        program = next((p for p in list_cnc_programs() if p["program_name"] == program_name), None)
        if not program:
            return
        rev = self.analysis_rev_var.get()
        revisions = list_cnc_program_revisions(program["id"])
        match = next((r for r in revisions if str(r["revision"]) == rev), None)
        if match:
            self.current_revision_id = match["id"]
            self.current_lines = self._read_program_lines(match["file_path"])
            self._update_analysis_view()
            self._refresh_code_panel(match["id"])
            self._refresh_observed_codes(match["id"])
            self._refresh_analysis_history()

    def _update_analysis_view(self):
        self.analysis_view.delete("1.0", tk.END)
        for idx, line in enumerate(self.current_lines, start=1):
            self.analysis_view.insert(tk.END, f"{idx:>5}  {line}")

    def run_analysis(self):
        if not self.current_revision_id:
            messagebox.showwarning("Select", "Select a program and revision first.")
            return
        lines = self.current_lines
        machine_profile = self.machine_profile_var.get().strip()
        result = analyze_program(lines)

        efficiency = result["efficiency_score"]
        cycle_time = result["cycle_time_seconds"]
        findings = result["findings"]
        tools = result["tools"]

        run_id = add_cnc_analysis_run(
            self.current_revision_id,
            machine_profile,
            efficiency,
            cycle_time,
        )
        for f in findings:
            add_cnc_finding(
                run_id,
                f.severity,
                f.rule_id,
                ",".join(str(x) for x in f.line_numbers),
                f.message,
                f.impact_seconds,
            )

        self.current_findings = findings
        self.current_tools = tools
        self.current_efficiency = efficiency
        self.current_cycle_time = cycle_time
        self.current_breakdown = result["breakdown"]
        self._render_analysis_summary(efficiency, cycle_time, findings)
        self._refresh_findings_list(findings)
        self._render_tool_summary(tools)
        self._refresh_analysis_history()

    def _render_analysis_summary(self, efficiency: float, cycle_time: float, findings: List[Finding]):
        self.analysis_summary.delete("1.0", tk.END)
        breakdown = getattr(self, "current_breakdown", {})
        self.analysis_summary.insert(
            tk.END,
            f"Efficiency Score: {efficiency:.1f}\nCycle Time Estimate: {cycle_time:.1f} sec\nFindings: {len(findings)}\n",
        )
        if breakdown:
            self.analysis_summary.insert(
                tk.END,
                f"Total Cut Distance (mm): {breakdown.get('cut_distance_mm', 0.0):.1f}\n"
                f"Total Move Distance (mm): {breakdown.get('move_distance_mm', 0.0):.1f}\n"
                f"Average Move (mm): {breakdown.get('avg_move_mm', 0.0):.2f}\n",
            )

    def _refresh_findings_list(self, findings: List[Finding]):
        self.findings_list.delete(0, "end")
        for f in findings:
            self.findings_list.insert("end", f"{f.severity} | {f.rule_id} | {f.message}")

    def _select_finding(self, event=None):
        sel = self.findings_list.curselection()
        if not sel:
            return
        finding = self.current_findings[sel[0]]
        if finding.line_numbers:
            line = finding.line_numbers[0]
            self.analysis_view.see(f"{line}.0")
            self.analysis_view.tag_remove("highlight", "1.0", tk.END)
            self.analysis_view.tag_add("highlight", f"{line}.0", f"{line}.0 lineend")
            self.analysis_view.tag_config("highlight", background="yellow")

    def _refresh_analysis_history(self):
        if not self.current_revision_id:
            return
        self.analysis_history = list_cnc_analysis_runs(self.current_revision_id)
        values = [
            f"{row['id']} | {row['machine_profile']} | {row['created_at']}"
            for row in self.analysis_history
        ]
        self.history_cb["values"] = values
        if values and not self.history_cb.get():
            self.history_cb.set(values[0])

    def _load_history_run(self, event=None):
        sel = self.history_cb.get()
        if not sel:
            return
        run_id = int(sel.split("|")[0].strip())
        run = next((r for r in self.analysis_history if r["id"] == run_id), None)
        findings = list_cnc_findings(run_id)
        self.current_findings = [
            Finding(
                f["severity"],
                f["rule_id"],
                [int(x) for x in f["line_numbers"].split(",") if x],
                f["message"],
                float(f.get("impact_seconds", 0.0)),
            )
            for f in findings
        ]
        self._refresh_findings_list(self.current_findings)
        if run:
            self.current_efficiency = float(run.get("efficiency_score", 0.0))
            self.current_cycle_time = float(run.get("cycle_time_seconds", 0.0))
            self.current_breakdown = {}
            self._render_analysis_summary(self.current_efficiency, self.current_cycle_time, self.current_findings)

    # ------------------- Tool Summary -------------------
    def _build_tool_summary(self):
        self.tool_tree = ttk.Treeview(
            self.tab_tool,
            columns=("tool", "calls", "cut_time"),
            show="headings",
            height=14,
        )
        self.tool_tree.heading("tool", text="TOOL")
        self.tool_tree.heading("calls", text="CALLS")
        self.tool_tree.heading("cut_time", text="CUT TIME (SEC)")
        self.tool_tree.pack(fill="both", expand=True, padx=10, pady=10)

    def _render_tool_summary(self, tools: Dict[str, Dict[str, float]]):
        for i in self.tool_tree.get_children():
            self.tool_tree.delete(i)
        for tool, info in sorted(tools.items()):
            self.tool_tree.insert("", "end", values=(
                tool,
                int(info.get("calls", 0)),
                float(info.get("cut_time", 0.0)),
            ))

    # ------------------- Reports -------------------
    def _build_reports(self):
        wrap = tk.Frame(self.tab_reports, padx=10, pady=10)
        wrap.pack(fill="both", expand=True)

        tk.Button(wrap, text="Export to Excel", command=self.export_excel).pack(anchor="w")
        tk.Button(wrap, text="Export to PDF", command=self.export_pdf).pack(anchor="w", pady=6)
        self.export_status = tk.Label(wrap, text="")
        self.export_status.pack(anchor="w", pady=10)

    def export_excel(self):
        if openpyxl is None:
            messagebox.showerror("Missing Dependency", "openpyxl is not available.")
            return
        path = os.path.join(CNC_EXPORTS_DIR, f"cnc_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Efficiency Score", self.current_efficiency])
        ws.append(["Cycle Time (sec)", self.current_cycle_time])
        breakdown = getattr(self, "current_breakdown", {})
        if breakdown:
            ws.append(["Total Cut Distance (mm)", breakdown.get("cut_distance_mm", 0.0)])
            ws.append(["Total Move Distance (mm)", breakdown.get("move_distance_mm", 0.0)])
            ws.append(["Average Move (mm)", breakdown.get("avg_move_mm", 0.0)])

        ws_findings = wb.create_sheet("Findings")
        ws_findings.append(["Severity", "Rule", "Lines", "Message", "Impact (sec)"])
        for f in self.current_findings:
            ws_findings.append([f.severity, f.rule_id, ",".join(map(str, f.line_numbers)), f.message, f.impact_seconds])

        ws_tools = wb.create_sheet("Tool Summary")
        ws_tools.append(["Tool", "Calls", "Cut Time (sec)"])
        for tool, info in self.current_tools.items():
            ws_tools.append([tool, int(info.get("calls", 0)), float(info.get("cut_time", 0.0))])

        wb.save(path)
        self.export_status.config(text=f"Saved Excel: {path}")

    def export_pdf(self):
        if canvas is None or letter is None:
            messagebox.showerror("Missing Dependency", "reportlab is not available.")
            return
        path = os.path.join(CNC_EXPORTS_DIR, f"cnc_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        c = canvas.Canvas(path, pagesize=letter)
        width, height = letter
        y = height - 40
        c.drawString(40, y, "CNC Analyzer Report")
        y -= 20
        c.drawString(40, y, f"Efficiency Score: {self.current_efficiency:.1f}")
        y -= 14
        c.drawString(40, y, f"Cycle Time (sec): {self.current_cycle_time:.1f}")
        y -= 20
        breakdown = getattr(self, "current_breakdown", {})
        if breakdown:
            c.drawString(40, y, f"Total Cut Distance (mm): {breakdown.get('cut_distance_mm', 0.0):.1f}")
            y -= 14
            c.drawString(40, y, f"Total Move Distance (mm): {breakdown.get('move_distance_mm', 0.0):.1f}")
            y -= 14
            c.drawString(40, y, f"Average Move (mm): {breakdown.get('avg_move_mm', 0.0):.2f}")
            y -= 20
        for f in self.current_findings:
            c.drawString(40, y, f"{f.severity} {f.rule_id}: {f.message}")
            y -= 14
            if y < 60:
                c.showPage()
                y = height - 40
        y -= 10
        c.drawString(40, y, "Tool Summary")
        y -= 16
        for tool, info in self.current_tools.items():
            c.drawString(40, y, f"T{tool} | Calls: {int(info.get('calls', 0))} | Cut Time: {info.get('cut_time', 0.0):.1f}")
            y -= 14
            if y < 60:
                c.showPage()
                y = height - 40
        c.save()
        self.export_status.config(text=f"Saved PDF: {path}")


def analyze_program(lines: List[str]) -> Dict[str, object]:
    units = "mm"
    feed_mode = "G94"
    current_tool = None
    spindle_on = False
    last_modal = None
    tool_calls = {}
    tool_sections = {}
    moves = []
    cut_time = {}
    findings: List[Finding] = []
    g_codes_seen = set()
    m_codes_seen = set()
    canned_active = False
    last_spindle_line = None
    retract_cycles = 0
    last_z = None
    unsupported_codes = set()
    feed_mode_flagged = False

    position = {"X": 0.0, "Y": 0.0, "Z": 0.0}
    move_lengths = []

    for idx, raw in enumerate(lines, start=1):
        line = raw.upper()
        codes = re.findall(r"\b([GM]\d+)\b", line)
        for code in codes:
            if code.startswith("G"):
                g_codes_seen.add(code)
            if code.startswith("M"):
                m_codes_seen.add(code)

        if "G20" in line:
            units = "inch"
        if "G21" in line:
            units = "mm"
        if "G95" in line:
            feed_mode = "G95"
        if "G94" in line:
            feed_mode = "G94"

        if re.search(r"\bM0?3\b", line) or re.search(r"\bM0?4\b", line):
            spindle_on = True
            last_spindle_line = idx
        if re.search(r"\bM0?5\b", line):
            spindle_on = False

        tool_match = re.search(r"\bT(\d+)\b", line)
        if tool_match:
            tool = tool_match.group(1)
            current_tool = tool
            tool_calls[tool] = tool_calls.get(tool, 0) + 1
            tool_sections.setdefault(tool, []).append(idx)

        if "G80" in line:
            canned_active = False
        if re.search(r"\bG8\d\b", line) and "G80" not in line:
            canned_active = True

        modal_match = re.search(r"\bG0?([0123])\b", line)
        if modal_match:
            modal = modal_match.group(0)
            if modal == last_modal:
                findings.append(Finding("INFO", "MODAL_SPAM", [idx], f"Redundant modal {modal} repeated."))
            last_modal = modal

        coords = {}
        for axis in ("X", "Y", "Z"):
            match = re.search(rf"{axis}([+-]?\d+(\.\d+)?)", line)
            if match:
                coords[axis] = float(match.group(1))

        if coords:
            target = position.copy()
            target.update(coords)
            dist = ((target["X"] - position["X"]) ** 2 + (target["Y"] - position["Y"]) ** 2 + (target["Z"] - position["Z"]) ** 2) ** 0.5
            if units == "inch":
                dist *= 25.4
            move_lengths.append(dist)
            moves.append((idx, dist, line))
            if last_z is not None and "Z" in coords:
                if coords["Z"] > last_z:
                    retract_cycles += 0.5
                if coords["Z"] < last_z:
                    retract_cycles += 0.5
            last_z = coords.get("Z", last_z)
            position = target
            if modal_match and modal_match.group(0) in ("G1", "G01", "G2", "G02", "G3", "G03"):
                if not spindle_on:
                    findings.append(Finding("CRITICAL", "SPINDLE_MISSING", [idx], "Cutting move without spindle start."))
                if current_tool:
                    cut_time[current_tool] = cut_time.get(current_tool, 0.0) + dist

        if feed_mode == "G95" and not feed_mode_flagged:
            findings.append(Finding("INFO", "FEED_UNSUPPORTED", [idx], "G95 feed mode detected."))
            feed_mode_flagged = True

    if canned_active:
        findings.append(Finding("CRITICAL", "CANNED_MISSING_G80", [], "Canned cycle used without G80 cancel."))
    if retract_cycles >= 6:
        findings.append(Finding("WARN", "EXCESS_RETRACTS", [], "Excessive retract cycles detected."))

    for tool, sections in tool_sections.items():
        if len(sections) > 1:
            findings.append(Finding("WARN", "TOOL_RECALL", sections, f"Tool {tool} recalled in multiple sections."))

    if move_lengths:
        avg_move = sum(move_lengths) / len(move_lengths)
        if avg_move < 0.5:
            findings.append(Finding("WARN", "TINY_SEGMENTS", [], "Average move length is very small."))

    for idx, dist, line in moves:
        if dist > 200 and "G1" in line:
            findings.append(Finding("INFO", "AIR_CUT", [idx], "Long cutting feed move may be air-cutting.", 1.0))

    supported_codes = {
        "G0", "G00", "G1", "G01", "G2", "G02", "G3", "G03",
        "G20", "G21", "G80", "G81", "G82", "G83", "G84", "G85", "G86", "G87", "G88", "G89",
        "G90", "G91", "G94", "G95",
        "M3", "M03", "M4", "M04", "M5", "M05", "M6", "M06", "M30",
    }
    for idx, raw in enumerate(lines, start=1):
        for code in re.findall(r"\b([GM]\d+)\b", raw.upper()):
            if code not in supported_codes and code not in unsupported_codes:
                unsupported_codes.add(code)
                findings.append(
                    Finding("INFO", "UNKNOWN_CODE", [idx], f"Unknown/Unhandled code encountered: {code}.")
                )

    tool_summary = {}
    for tool, count in tool_calls.items():
        tool_summary[tool] = {"calls": count, "cut_time": cut_time.get(tool, 0.0)}

    cycle_time = sum(cut_time.values()) / 10.0 if cut_time else 0.0
    efficiency = max(0.0, 100.0 - len(findings) * 2.5)
    if feed_mode == "G95":
        efficiency = max(0.0, efficiency - 10.0)

    return {
        "efficiency_score": efficiency,
        "cycle_time_seconds": cycle_time,
        "findings": findings,
        "tools": tool_summary,
        "breakdown": {
            "cut_distance_mm": float(sum(cut_time.values())),
            "move_distance_mm": float(sum(move_lengths)),
            "avg_move_mm": float(sum(move_lengths) / len(move_lengths)) if move_lengths else 0.0,
        },
    }
